"""
==========================================================
HELIO-FORGE (SOLAR PRELUDE)
dataset_exporter.py

Export datasets to CSV, Parquet and Excel.
==========================================================
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import platform

import pandas as pd

from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Font,
    Alignment,
)

from openpyxl.utils import get_column_letter

from src.utils.config import PATH_CFG

# Export filenames — match data_paths.yaml features section
_CSV_FILENAME     = "selected_features.csv"
_PARQUET_FILENAME = "selected_features.parquet"
_EXCEL_FILENAME   = "selected_features.xlsx"


class DatasetExporter:
    """
    Export datasets into multiple formats.

    Supported formats
    -----------------
    • CSV
    • Parquet
    • Excel (.xlsx)
    """

    def __init__(
        self,
        output_directory: str | Path,
    ) -> None:

        self.output_directory = (
            Path(output_directory)
            if output_directory is not None
            else PATH_CFG.features.root
        )

        self.output_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        ##################################################
        # STYLES
        ##################################################

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        self.header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        self.center = Alignment(
            horizontal="center",
            vertical="center",
        )

        self.border = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin"),

        )

    ##################################################
    # PRIVATE HELPERS
    ##################################################

    def _style_sheet(
        self,
        worksheet,
    ) -> None:
        """
        Apply common worksheet styling.
        """

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.alignment = self.center

            cell.border = self.border

        for row in worksheet.iter_rows():

            for cell in row:

                cell.border = self.border

        ##################################################
        # AUTO WIDTH
        ##################################################

        for column in worksheet.columns:

            length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    length = max(
                        length,
                        len(str(cell.value)),
                    )

                except Exception:
                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                length + 3,
                40,
            )
            
    ##################################################
    # CSV EXPORT
    ##################################################

    def export_csv(
        self,
        dataframe: pd.DataFrame,
        filename: str | None = None,
    ) -> Path:
        """
        Export dataset to CSV.
        """

        resolved_filename = filename or _CSV_FILENAME
        output_file = self.output_directory / resolved_filename

        dataframe.to_csv(
            output_file,
            index=False,
        )

        return output_file

    ##################################################
    # PARQUET EXPORT
    ##################################################

    def export_parquet(
        self,
        dataframe: pd.DataFrame,
        filename: str | None = None,
    ) -> Path:
        """
        Export dataset to Parquet.
        """

        resolved_filename = filename or _PARQUET_FILENAME
        output_file = self.output_directory / resolved_filename

        dataframe.to_parquet(
            output_file,
            index=False,
        )

        return output_file

    ##################################################
    # EXCEL EXPORT
    ##################################################

    def export_excel(
        self,
        dataframe: pd.DataFrame,
        filename: str | None = None,
    ) -> Path:
        """
        Export dataset as a professional Excel workbook.
        """

        resolved_filename = filename or _EXCEL_FILENAME
        output_file = self.output_directory / resolved_filename

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
        ) as writer:

            ##################################################
            # DATASET SUMMARY
            ##################################################

            summary = pd.DataFrame({

                "Property": [

                    "Samples",

                    "Features",

                    "Missing Values",

                    "Duplicate Rows",

                    "Memory Usage (MB)",

                    "Generated",

                ],

                "Value": [

                    len(dataframe),

                    len(dataframe.columns),

                    int(
                        dataframe.isna().sum().sum()
                    ),

                    int(
                        dataframe.duplicated().sum()
                    ),

                    round(

                        dataframe.memory_usage(
                            deep=True,
                        ).sum()

                        / (1024 ** 2),

                        2,

                    ),

                    datetime.now().strftime(
                        "%d-%m-%Y %H:%M:%S"
                    ),

                ],

            })

            summary.to_excel(

                writer,

                sheet_name="Dataset Summary",

                index=False,

            )

            ws = writer.sheets[
                "Dataset Summary"
            ]

            self._style_sheet(ws)
            
            ##################################################
            # ALL FEATURES
            ##################################################

            dataframe.to_excel(
                writer,
                sheet_name="All Features",
                index=False,
            )

            ws = writer.sheets[
                "All Features"
            ]

            self._style_sheet(ws)

            ##################################################
            # FEATURE STATISTICS
            ##################################################

            statistics = (
                dataframe.describe(
                    include="all",
                )
                .transpose()
                .reset_index()
                .rename(
                    columns={
                        "index": "Feature",
                    }
                )
            )

            statistics.to_excel(
                writer,
                sheet_name="Feature Statistics",
                index=False,
            )

            ws = writer.sheets[
                "Feature Statistics"
            ]

            self._style_sheet(ws)

            ##################################################
            # MISSING VALUES
            ##################################################

            missing = pd.DataFrame({

                "Feature": dataframe.columns,

                "Missing Values": dataframe
                .isna()
                .sum()
                .values,

                "Missing Percentage": (

                    dataframe
                    .isna()
                    .mean()
                    .values

                    * 100

                ).round(2),

            })

            missing.to_excel(
                writer,
                sheet_name="Missing Values",
                index=False,
            )

            ws = writer.sheets[
                "Missing Values"
            ]

            self._style_sheet(ws)
            
                ##################################################
            # CORRELATION MATRIX
            ##################################################

            correlation = dataframe.corr(
                numeric_only=True,
            )

            correlation.to_excel(
                writer,
                sheet_name="Correlation Matrix",
            )

            ws = writer.sheets[
                "Correlation Matrix"
            ]

            self._style_sheet(ws)

            ##################################################
            # METADATA
            ##################################################

            metadata = pd.DataFrame({

                "Property": [

                    "Rows",

                    "Columns",

                    "Generated",

                    "Python Version",

                    "Pandas Version",

                ],

                "Value": [

                    len(dataframe),

                    len(dataframe.columns),

                    datetime.now().strftime(
                        "%d-%m-%Y %H:%M:%S"
                    ),

                    platform.python_version(),

                    pd.__version__,

                ],

            })

            metadata.to_excel(
                writer,
                sheet_name="Metadata",
                index=False,
            )

            ws = writer.sheets[
                "Metadata"
            ]

            self._style_sheet(ws)

        return output_file

    ##################################################
    # EXPORT EVERYTHING
    ##################################################

    def export_all(
        self,
        dataframe: pd.DataFrame,
    ) -> dict[str, Path]:
        """
        Export dataset to every supported format.

        Returns
        -------
        dict[str, Path]
            Paths of generated files.
        """

        csv_path = self.export_csv(
            dataframe,
        )

        parquet_path = self.export_parquet(
            dataframe,
        )

        excel_path = self.export_excel(
            dataframe,
        )

        return {

            "csv": csv_path,

            "parquet": parquet_path,

            "excel": excel_path,

        }
        
#####################
        #TODO:
            #- Improve features.xlsx to include observation-wise engineering report
            #- Pair all SoLEXS and HEL1OS observations automatically
            #- Add observation metadata sheet
#####################